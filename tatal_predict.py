import os
import torch
from torch import nn
from torchvision import transforms
from model import RDR_model
from PIL import Image
from unit import to_psnr, to_ssim_skimage, to_ssim, to_pearson
import openpyxl


wb = openpyxl.Workbook()
ws = wb.create_sheet('2f')
ws.cell(row=1, column=1).value = "mse"
ws.cell(row=1, column=2).value = "psnr"
ws.cell(row=1, column=3).value = "npcc"
ws.cell(row=1, column=4).value = "ssim"

Tensor = transforms.Compose([transforms.Resize((256, 256)), transforms.ToTensor()])
RDR_model = RDR_model()
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")  
RDR_model.to(device)
RDR_model.load_state_dict(torch.load('./parameter/mse_gr36_f'))  
print("model ready")
img_path = 'D:/1'
lb_path = 'D:/2'

imglist = os.listdir(img_path)
imglist.sort(key=lambda x: int(x[:-4]))
s = 1
for i in imglist:
    image_path = os.path.join(img_path, i)
    label_path = os.path.join(lb_path, i)
    image = Image.open(image_path)
    image = Tensor(image)
    image = torch.reshape(image, (1, 1, 256, 256))
    label = Image.open(label_path)
    label = Tensor(label)
    label = torch.reshape(label, (1, 1, 256, 256))
    b = torch.ones(256, 256).to(device)

    RDR_model.eval()
    with torch.no_grad():
        output = image.to(device)
        label = label.to(device)
        # output = RDR_model(img)
        # output = torch.where(output < 1, output, b)
        # print(torch.max(output))
        mse_loss = nn.MSELoss()
        mse = mse_loss(output, label)
        # print(mse)
        psnr = to_psnr(output, label)
        npcc = to_pearson(output, label)
        # print(psnr)
        # print(npcc)
        ssim = to_ssim_skimage(output, label)
        # ssim1 = to_ssim(output, label)
        # print(ssim)
        # print(ssim1)
        s += 1
        ws.cell(row=s, column=1).value = mse.item()
        ws.cell(row=s, column=2).value = psnr
        ws.cell(row=s, column=3).value = npcc.item()
        ws.cell(row=s, column=4).value = ssim[0]
        wb.save('2f.xlsx')
